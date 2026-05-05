# server.py
# -------------------------------------------------------
# Flask REST API — 후보기술 개요서 자동 요약 봇 백엔드
# -------------------------------------------------------

import os
import re
import io
import json
import uuid
import time
import sys
import threading
import tempfile
from pathlib import Path
from datetime import datetime
# import appnope

# ============================================================
# Project root resolution (backend/ 폴더에서 실행해도 프로젝트 루트 기준 동작)
# ============================================================
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(Path(__file__).resolve().parent))  # backend/ 내 모듈 import용

from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS

try:
    from dotenv import load_dotenv
    load_dotenv(PROJECT_ROOT / ".env")
    print("ENV PATH:", PROJECT_ROOT / ".env")
    print("ENV EXISTS:", (PROJECT_ROOT / ".env").exists())
    print("OPENAI KEY:", os.getenv("OPENAI_API_KEY"))
except ImportError:
    pass

# 로컬 모듈 (backend/ 내)
from metadata_extractor import DocMeta, extract_pdf_metadata, extract_patent_from_url, scan_input_folder
from matcher import RefItem, MatchResult, load_reference_list, match_documents, results_to_dataframe
from number_converter import load_mapping_table, apply_conversion, save_conversion_state
from ppt_manager import reorder_by_number_mapping, get_pptx_info

FRONTEND_DIST = PROJECT_ROOT / "frontend" / "dist"

app = Flask(__name__, static_folder=str(FRONTEND_DIST), static_url_path="/")
CORS(app)

# ============================================================
# Serve React Frontend
# ============================================================
@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve_frontend(path):
    if path != "" and (FRONTEND_DIST / path).exists() and (FRONTEND_DIST / path).is_file():
        return send_from_directory(app.static_folder, path)
    else:
        return send_from_directory(app.static_folder, 'index.html')

# ============================================================
# In-memory session store (single-user local usage)
# ============================================================
DEFAULT_TEMPLATE = str(PROJECT_ROOT / "후보기술 개요서 템플릿_선택창 이름 생성_v2.pptx")
DEFAULT_OUTPUT_DIR = str(PROJECT_ROOT / "output")

# .env에서 API Key 자동 로드
_ENV_OPENAI_KEY = os.getenv("OPENAI_API_KEY", "").strip()
_ENV_GEMINI_KEY = os.getenv("GEMINI_API_KEY", "").strip()

session_store = {
    "documents": [],        # List[dict] — scanned document metadata
    "order_data": [],       # List[dict] — order excel data
    "ref_list": [],         # List[RefItem]
    "match_results": [],    # List[dict] — matching preview
    "jobs": {},             # job_id -> job state
    "api_key": _ENV_OPENAI_KEY,
    "gemini_api_key": _ENV_GEMINI_KEY,
    "model": "gpt-5",
    "template_path": DEFAULT_TEMPLATE,
    "output_dir": DEFAULT_OUTPUT_DIR,
}

OUTPUT_DIR = Path(DEFAULT_OUTPUT_DIR)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


# ============================================================
# Helpers
# ============================================================
def _doc_meta_to_dict(dm: DocMeta) -> dict:
    return {
        "filename": dm.filename,
        "doc_type": dm.doc_type,
        "title": dm.title or "(추출 실패)",
        "doi": dm.doi or "",
        "patent_no": dm.patent_no or "",
        "institution": dm.institution or "",
        "year": dm.year or "",
        "source_type": dm.source_type,
        "source_path": dm.source_path,
        "has_text": bool(dm.full_text and len(dm.full_text) > 100),
        "figure_count": len(dm.figure_urls),
    }


def _store_doc_meta(dm: DocMeta):
    """Store DocMeta both as dict (for API) and keep the object for processing."""
    d = _doc_meta_to_dict(dm)
    d["_full_text"] = dm.full_text  # keep internally
    d["_figure_urls"] = dm.figure_urls  # keep internally
    return d


# ============================================================
# API Endpoints
# ============================================================

# ---------- Settings ----------
@app.route("/api/settings", methods=["GET"])
def get_settings():
    return jsonify({
        "api_key_set": bool(session_store["api_key"]),
        "gemini_key_set": bool(session_store.get("gemini_api_key")),
        "model": session_store["model"],
        "template_path": session_store["template_path"],
        "output_dir": session_store["output_dir"],
        "document_count": len(session_store["documents"]),
    })


@app.route("/api/settings", methods=["POST"])
def update_settings():
    data = request.json or {}
    if "api_key" in data:
        session_store["api_key"] = data["api_key"]
        os.environ["OPENAI_API_KEY"] = data["api_key"]
    if "gemini_api_key" in data:
        session_store["gemini_api_key"] = data["gemini_api_key"]
        os.environ["GEMINI_API_KEY"] = data["gemini_api_key"]
    if "model" in data:
        session_store["model"] = data["model"]
    if "template_path" in data:
        session_store["template_path"] = data["template_path"]
    if "output_dir" in data:
        session_store["output_dir"] = data["output_dir"]
        Path(data["output_dir"]).mkdir(parents=True, exist_ok=True)
    return jsonify({"ok": True})


# ---------- Document Input: PDF Upload ----------
@app.route("/api/upload-pdfs", methods=["POST"])
def upload_pdfs():
    """Upload multiple PDF files, extract metadata, store in session.
    reset=true 쿼리 파라미터가 있으면 기존 문서 리스트를 먼저 초기화.
    """
    # reset 플래그: 요약 완료 후 다시 실행 시 젬 문서 제거
    if request.form.get("reset", "").lower() == "true":
        session_store["documents"] = []
        session_store["match_results"] = []
        session_store["order_data"] = []

    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "No files provided"}), 400

    results = []
    for f in files:
        if not f.filename.lower().endswith(".pdf"):
            continue
        # Save to temp
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            f.save(tmp)
            tmp_path = tmp.name
        try:
            dm = extract_pdf_metadata(tmp_path)
            dm.filename = f.filename
            dm.source_path = tmp_path  # keep temp path for later processing
            d = _store_doc_meta(dm)
            session_store["documents"].append(d)
            results.append(_doc_meta_to_dict(dm))
        except Exception as e:
            results.append({"filename": f.filename, "error": str(e)})

    return jsonify({
        "uploaded": len(results),
        "documents": results,
        "total": len(session_store["documents"]),
    })


# ---------- Document Input: Local Folder Scan ----------
@app.route("/api/scan-folder", methods=["POST"])
def scan_folder():
    """Scan a local folder for PDF files."""
    data = request.json or {}
    folder_path = data.get("folder_path", "")
    if not folder_path or not os.path.isdir(folder_path):
        return jsonify({"error": f"Invalid folder path: {folder_path}"}), 400

    # reset 플래그
    if data.get("reset", False):
        session_store["documents"] = []
        session_store["match_results"] = []
        session_store["order_data"] = []

    try:
        metas = scan_input_folder(folder_path)
        results = []
        for dm in metas:
            d = _store_doc_meta(dm)
            session_store["documents"].append(d)
            results.append(_doc_meta_to_dict(dm))
        return jsonify({
            "scanned": len(results),
            "documents": results,
            "total": len(session_store["documents"]),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ---------- Document Input: Patent Excel Upload ----------
@app.route("/api/upload-patent-excel", methods=["POST"])
def upload_patent_excel():
    """Upload Excel with patent URLs, scrape each URL."""
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file provided"}), 400

    import pandas as pd

    with tempfile.NamedTemporaryFile(suffix=f".{f.filename.split('.')[-1]}", delete=False) as tmp:
        f.save(tmp)
        tmp_path = tmp.name

    try:
        if tmp_path.endswith(".csv"):
            df = pd.read_csv(tmp_path, encoding="utf-8-sig")
        else:
            df = pd.read_excel(tmp_path, engine="openpyxl")

        # Find URL column
        url_col = None
        for c in df.columns:
            cl = str(c).lower()
            if any(kw in cl for kw in ["url", "링크", "link", "주소", "wips"]):
                url_col = c
                break
        if url_col is None:
            # Try first column containing URLs
            for c in df.columns:
                sample = str(df[c].iloc[0]) if len(df) > 0 else ""
                if "http" in sample:
                    url_col = c
                    break

        if url_col is None:
            return jsonify({"error": "URL 컬럼을 찾을 수 없습니다"}), 400

        # reset 플래그
        if request.form.get("reset", "").lower() == "true":
            session_store["documents"] = []
            session_store["match_results"] = []
            session_store["order_data"] = []

        results = []
        errors = []
        for idx, row in df.iterrows():
            url = str(row[url_col]).strip()
            if not url.startswith("http"):
                continue
            try:
                dm = extract_patent_from_url(url)
                d = _store_doc_meta(dm)
                session_store["documents"].append(d)
                results.append(_doc_meta_to_dict(dm))
            except Exception as e:
                errors.append({"url": url, "error": str(e)})

        return jsonify({
            "processed": len(results),
            "errors": errors,
            "documents": results,
            "total": len(session_store["documents"]),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        os.unlink(tmp_path)


# ---------- Documents List ----------
@app.route("/api/documents", methods=["GET"])
def list_documents():
    docs = [{k: v for k, v in d.items() if not k.startswith("_")}
            for d in session_store["documents"]]
    return jsonify({"documents": docs, "total": len(docs)})


@app.route("/api/documents", methods=["DELETE"])
def clear_documents():
    session_store["documents"] = []
    session_store["match_results"] = []
    session_store["order_data"] = []
    return jsonify({"ok": True})


@app.route("/api/documents/<int:index>", methods=["DELETE"])
def delete_document(index):
    if 0 <= index < len(session_store["documents"]):
        session_store["documents"].pop(index)
        return jsonify({"ok": True, "total": len(session_store["documents"])})
    return jsonify({"error": "Index out of range"}), 400


# ---------- Order Excel Upload & Matching ----------
@app.route("/api/upload-order-excel", methods=["POST"])
def upload_order_excel():
    """Upload ordering Excel, match with scanned documents."""
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file provided"}), 400

    import pandas as pd

    index_col = request.form.get("index_col")

    with tempfile.NamedTemporaryFile(suffix=f".{f.filename.split('.')[-1]}", delete=False) as tmp:
        f.save(tmp)
        tmp_path = tmp.name

    try:
        ref_list = load_reference_list(tmp_path, index_col=index_col if index_col else None)
        session_store["ref_list"] = ref_list

        # Build DocMeta objects for matching
        doc_metas = []
        for d in session_store["documents"]:
            dm = DocMeta(
                filename=d["filename"],
                doc_type=d["doc_type"],
                title=d.get("title", ""),
                doi=d.get("doi", ""),
                patent_no=d.get("patent_no", ""),
                source_type=d.get("source_type", "pdf"),
                source_path=d.get("source_path", ""),
            )
            doc_metas.append(dm)

        # Run matching
        match_results = match_documents(doc_metas, ref_list)

        # Convert to serializable format
        match_data = []
        for mr in match_results:
            match_data.append({
                "filename": mr.doc_meta.filename,
                "doc_type": mr.doc_meta.doc_type,
                "title": mr.doc_meta.title[:80] if mr.doc_meta.title else "",
                "ref_index": mr.ref_item.index if mr.ref_item else None,
                "ref_title": mr.ref_item.title[:60] if mr.ref_item else "",
                "score": round(mr.score, 1),
                "match_type": mr.match_type,
                "needs_review": mr.needs_review,
                "status": mr.status_label,
            })

        session_store["match_results"] = match_data
        session_store["order_data"] = [
            {"index": r.index, "title": r.title, "category": r.category}
            for r in ref_list
        ]

        return jsonify({
            "ref_count": len(ref_list),
            "match_results": match_data,
            "order_data": session_store["order_data"],
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        os.unlink(tmp_path)


# ---------- Summarize ----------
@app.route("/api/summarize", methods=["POST"])
def start_summarize():
    """Kick off summarization in background thread."""
    data = request.json or {}
    api_key = data.get("api_key", session_store.get("api_key", ""))
    model = data.get("model", session_store.get("model", "gpt-4o"))
    template = data.get("template_path", session_store.get("template_path", ""))
    out_dir = data.get("output_dir", session_store.get("output_dir", "output"))

    if not api_key:
        return jsonify({"error": "API Key가 필요합니다"}), 400
    if not template or not os.path.exists(template):
        return jsonify({"error": f"템플릿 파일 없음: {template}"}), 400
    if not session_store["documents"]:
        return jsonify({"error": "문서가 없습니다. 먼저 문서를 입력하세요."}), 400

    os.environ["OPENAI_API_KEY"] = api_key
    session_store["api_key"] = api_key
    session_store["model"] = model

    job_id = str(uuid.uuid4())[:8]
    job = {
        "id": job_id,
        "status": "running",
        "progress": 0,
        "total": len(session_store["documents"]),
        "current_file": "",
        "completed": [],
        "errors": [],
        "output_file": None,
        "json_files": [],    # 현재 세션에서 생성된 JSON 파일 경로 목록
        "started_at": datetime.now().isoformat(),
    }
    session_store["jobs"][job_id] = job

    def _run_summarize():
        try:
            # 방해 금지 (맥OS 잠자기 방지)
            # appnope.nope()

            from build_briefs_v2 import (
                extract_text, extract_images, pick_two_figures_with_vision,
                process_one_document, fill_slide, duplicate_slide, call_llm,
            )
            from pptx import Presentation as Prs
            import fitz

            out_path = Path(out_dir)
            out_path.mkdir(parents=True, exist_ok=True)

            prs = Prs(template)
            base_slide_idx = 0
            # 템플릿 파일의 원래 슬라이드 수를 기억해 둠 (요약 후 전부 제거)
            template_slide_count = len(prs.slides)

            docs = list(session_store["documents"])

            # ── match_results가 있으면 ref_index 기준으로 정렬 & 번호 반영 ──
            match_map = {}  # filename -> match_result dict
            for mr in session_store.get("match_results", []):
                if mr.get("filename"):
                    match_map[mr["filename"]] = mr

            if match_map:
                # ref_index가 있는 문서는 해당 번호순, 없는 문서는 맨 뒤
                def _sort_key(d):
                    mr = match_map.get(d["filename"])
                    if mr and mr.get("ref_index") is not None:
                        return (0, mr["ref_index"])
                    return (1, 0)
                docs.sort(key=_sort_key)
                print(f"[INFO] 문서를 ref_index 기준으로 정렬: "
                      f"{[(d['filename'], match_map.get(d['filename'], {}).get('ref_index', '?')) for d in docs]}")

            # 미매칭 문서에 부여할 번호 결정 (매칭된 번호 이후)
            used_indices = {mr["ref_index"] for mr in match_map.values() if mr.get("ref_index") is not None}
            next_unmatched = (max(used_indices) + 1) if used_indices else 1

            for idx, d in enumerate(docs):
                job["progress"] = idx
                job["current_file"] = d["filename"]

                try:
                    text = d.get("_full_text", "")
                    concept_fig, effect_fig, vision_caps = None, None, None
                    precomputed_raw = None

                    # ── stem / doc_idx: ref_index 우선, 없으면 자동 증가 ──
                    mr = match_map.get(d["filename"])
                    if mr and mr.get("ref_index") is not None:
                        doc_idx = mr["ref_index"]
                        stem = str(doc_idx)
                    else:
                        doc_idx = next_unmatched
                        stem = str(doc_idx)
                        next_unmatched += 1

                    if not text and d.get("source_path") and d["source_type"] == "pdf":
                        with fitz.open(d["source_path"]) as doc:
                            text = extract_text(doc)
                            # Step 1: 텍스트 LLM 먼저 호출 → representative_figures 확보
                            precomputed_raw = call_llm(text, model=model)
                            # 논문만 fig_title 전달 — 특허는 Vision LLM이 독립 판단
                            _is_patent = precomputed_raw.get("doc_type") == "patent"
                            rep_figs = [] if _is_patent else (precomputed_raw.get("representative_figures") or [])
                            # Step 2: Vision LLM에 fig_title 전달
                            try:
                                candidates = extract_images(doc)
                                concept_fig, effect_fig, vision_caps = pick_two_figures_with_vision(
                                    candidates, text, model=model,
                                    representative_figures=rep_figs,
                                )
                            except Exception:
                                vision_caps = ["", ""]
                    elif text and d["source_type"] == "pdf" and d.get("source_path"):
                        # Step 1: 텍스트 LLM 먼저 호출 → representative_figures 확보
                        precomputed_raw = call_llm(text, model=model)
                        # 논문만 fig_title 전달 — 특허는 Vision LLM이 독립 판단
                        _is_patent = precomputed_raw.get("doc_type") == "patent"
                        rep_figs = [] if _is_patent else (precomputed_raw.get("representative_figures") or [])
                        # Step 2: Vision LLM에 fig_title 전달
                        try:
                            with fitz.open(d["source_path"]) as doc:
                                candidates = extract_images(doc)
                                concept_fig, effect_fig, vision_caps = pick_two_figures_with_vision(
                                    candidates, text, model=model,
                                    representative_figures=rep_figs,
                                )
                        except Exception:
                            vision_caps = ["", ""]
                    elif d["source_type"] == "url":
                        # URL 특허: 텍스트 LLM 먼저 호출
                        precomputed_raw = call_llm(text, model=model)
                        # URL 특허는 항상 특허 — Vision LLM이 독립 판단
                        rep_figs = []
                        # figure_urls에서 도면 이미지 다운로드
                        fig_urls = d.get("_figure_urls", [])
                        if fig_urls:
                            try:
                                import requests as _req
                                url_candidates = []
                                for seq_idx, fig_url in enumerate(fig_urls[:10]):  # 최대 10장
                                    try:
                                        img_resp = _req.get(
                                            fig_url, timeout=15,
                                            headers={"User-Agent": "Mozilla/5.0"},
                                        )
                                        if img_resp.status_code == 200:
                                            img_bytes = img_resp.content
                                            # jpg → png 변환 (PIL 사용 가능한 경우)
                                            try:
                                                from PIL import Image as _Img
                                                import io as _io2
                                                im = _Img.open(_io2.BytesIO(img_bytes)).convert("RGB")
                                                w2, h2 = im.width, im.height
                                                out_buf = _io2.BytesIO()
                                                im.save(out_buf, format="PNG")
                                                png_bytes = out_buf.getvalue()
                                                area2 = w2 * h2
                                            except Exception:
                                                # PIL 없으면 jpg bytes 그대로
                                                png_bytes = img_bytes
                                                area2 = 500 * 500
                                            url_candidates.append({
                                                "page": seq_idx + 1,
                                                "area": area2,
                                                "png_bytes": png_bytes,
                                                "caption_hint": f"특허 도면 {seq_idx + 1}",
                                            })
                                    except Exception:
                                        continue
                                if url_candidates:
                                    concept_fig, effect_fig, vision_caps = pick_two_figures_with_vision(
                                        url_candidates, text, model=model,
                                        representative_figures=rep_figs,
                                    )
                                else:
                                    vision_caps = ["", ""]
                            except Exception as e_fig:
                                print(f"[WARN] URL 도면 수집 실패: {e_fig}")
                                vision_caps = ["", ""]
                        else:
                            vision_caps = ["", ""]

                    if not text:
                        job["errors"].append({"filename": d["filename"], "error": "텍스트 없음"})
                        continue

                    # 원본 메타 정보를 JSON에 함께 저장
                    if precomputed_raw is not None:
                        precomputed_raw["_meta_doc_type"] = d.get("doc_type", "")
                        precomputed_raw["_meta_doi"] = d.get("doi", "")
                        precomputed_raw["_meta_patent_no"] = d.get("patent_no", "")
                        precomputed_raw["_meta_institution"] = d.get("institution", "")
                        precomputed_raw["_meta_year"] = d.get("year", "")

                    brief, cf, ef = process_one_document(
                        text=text, stem=stem, model=model,
                        save_json=True, save_md=True,
                        out_dir=out_path,
                        concept_fig=concept_fig, effect_fig=effect_fig,
                        vision_caps=vision_caps, precomputed_raw=precomputed_raw,
                        source_pdf=d["filename"],
                    )

                    slide = duplicate_slide(prs, slide_index=base_slide_idx)
                    fill_slide(slide, brief, doc_idx=doc_idx, concept_fig=cf, effect_fig=ef)
                    job["completed"].append(d["filename"])
                    # 이번 세션에서 생성된 JSON 파일 경로 관리
                    json_path = str(out_path / f"{stem}.json")
                    if json_path not in job["json_files"]:
                        job["json_files"].append(json_path)

                except Exception as e:
                    job["errors"].append({"filename": d["filename"], "error": str(e)})

            # 템플릿 원본 슬라이드 전부 제거 (소개/안내 페이지 포함)
            from lxml import etree
            from pptx.opc.constants import RELATIONSHIP_TYPE as RT
            slide_list = prs.slides._sldIdLst
            n_remove = template_slide_count  # 템플릿 파일에 있던 슬라이드 수만큼
            slide_ids_to_remove = list(slide_list)[:n_remove]
            for sldId in slide_ids_to_remove:
                rId = sldId.get("r:id")
                if rId:
                    try:
                        prs.part.drop_rel(rId)
                    except Exception:
                        pass
                slide_list.remove(sldId)

            # 고아 슬라이드 관계 정리 (ZIP 중복 방지)
            active_rids = set()
            for sldId in slide_list:
                rid = sldId.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id") or sldId.get("r:id")
                if rid:
                    active_rids.add(rid)
            for rId, rel in list(prs.part.rels.items()):
                if rel.reltype == RT.SLIDE and rId not in active_rids:
                    try:
                        prs.part.drop_rel(rId)
                    except Exception:
                        pass

            ppt_filename = "output_briefs.pptx"
            ppt_path = str(out_path / ppt_filename)
            prs.save(ppt_path)

            job["output_file"] = ppt_filename
            job["status"] = "done"
            job["progress"] = job["total"]

        except Exception as e:
            job["status"] = "error"
            job["errors"].append({"filename": "system", "error": str(e)})

        # finally:
            # 작업이 끝나면 다시 잠자기 허용
            # appnope.nap()

    thread = threading.Thread(target=_run_summarize, daemon=True)
    thread.start()

    return jsonify({"job_id": job_id, "status": "running"})


# ---------- Job Status ----------
@app.route("/api/status/<job_id>", methods=["GET"])
def job_status(job_id):
    job = session_store["jobs"].get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    return jsonify({
        "id": job["id"],
        "status": job["status"],
        "progress": job["progress"],
        "total": job["total"],
        "current_file": job["current_file"],
        "completed_count": len(job["completed"]),
        "error_count": len(job["errors"]),
        "errors": job["errors"],
        "output_file": job["output_file"],
    })


# ---------- Reorder ----------
@app.route("/api/reorder", methods=["POST"])
def reorder_ppt():
    """Reorder PPT slides based on uploaded order Excel.

    요약 완료 후 생성된 PPT의 슬라이드를 match_results 기반으로 재배치합니다.
    각 슬라이드의 레이블에서 실제 before 번호를 추출하여 매핑을 구성합니다.
    """
    out_dir = session_store.get("output_dir", "output")
    input_pptx = str(Path(out_dir) / "output_briefs.pptx")

    if not os.path.exists(input_pptx):
        return jsonify({"error": "요약 PPT가 없습니다. 먼저 요약을 실행하세요."}), 400

    match_results = session_store.get("match_results", [])
    order_data = session_store.get("order_data", [])

    if not order_data:
        return jsonify({"error": "순서 엑셀이 없습니다. 먼저 순서 엑셀을 업로드하세요."}), 400

    try:
        # PPT 슬라이드의 실제 before 번호 추출
        ppt_info = get_pptx_info(input_pptx)
        slide_numbers = []
        for info in ppt_info:
            num = info.get("number")
            if num is not None:
                slide_numbers.append(num)
            else:
                slide_numbers.append(info["index"] + 1)

        # match_results에서 {before번호: after번호(ref_index)} 매핑 구성
        # match_results와 슬라이드가 같은 순서라고 가정
        mapping = {}
        for i, mr in enumerate(match_results):
            if mr.get("ref_index") is not None and i < len(slide_numbers):
                before_num = slide_numbers[i]
                mapping[before_num] = mr["ref_index"]

        output_filename = "output_briefs_reordered.pptx"
        output_pptx = str(Path(out_dir) / output_filename)

        if mapping:
            reorder_by_number_mapping(input_pptx, mapping, output_pptx)
        else:
            # No mapping available, just copy
            import shutil
            shutil.copy2(input_pptx, output_pptx)

        return jsonify({
            "ok": True,
            "output_file": output_filename,
            "mapping_used": {str(k): v for k, v in mapping.items()},
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ---------- Analyze Excel Headers ----------
@app.route("/api/analyze-excel-headers", methods=["POST"])
def analyze_excel_headers():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file provided"}), 400

    import pandas as pd
    excel_ext = f.filename.split('.')[-1] if f.filename else 'xlsx'
    with tempfile.NamedTemporaryFile(suffix=f".{excel_ext}", delete=False) as tmp:
        f.save(tmp)
        tmp_path = tmp.name

    try:
        if excel_ext == 'csv':
            df = pd.read_csv(tmp_path, encoding='utf-8-sig', nrows=0)
        else:
            df = pd.read_excel(tmp_path, engine='openpyxl', nrows=0)

        cols = [str(c).strip() for c in df.columns]

        def _find_col(keywords):
            for kw in keywords:
                for c in cols:
                    if kw in c:
                        return c
            return ""

        recommended_new = _find_col(['개요서 번호', '개요서번호'])
        recommended_old = _find_col(['기존 번호', '기존번호', '이전 순번', '이전순번', '번호', 'No.'])

        return jsonify({
            "headers": cols,
            "recommended_old_col": recommended_old,
            "recommended_new_col": recommended_new
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        os.unlink(tmp_path)



# ---------- Reorder Existing (Case B: PPT + Excel 직접 업로드) ----------
@app.route("/api/reorder-existing", methods=["POST"])
def reorder_existing_ppt():
    """Case B: 이미 생성된 PPT 파일과 순번 엑셀을 업로드하여 재정렬.

    엑셀에서 {before번호(기존번호): after번호(개요서번호)} 매핑을 읽어
    PPT 슬라이드를 after 번호 순서로 재배치하고
    각 슬라이드의 레이블 번호도 after 번호로 변경합니다.
    """
    ppt_file = request.files.get("ppt_file")
    excel_file = request.files.get("excel_file")
    if not ppt_file or not excel_file:
        return jsonify({"error": "PPT 파일과 엑셀 파일 모두 필요합니다"}), 400

    # 임시 파일 저장
    with tempfile.NamedTemporaryFile(suffix=".pptx", delete=False) as tmp_ppt:
        ppt_file.save(tmp_ppt)
        tmp_ppt_path = tmp_ppt.name

    excel_ext = excel_file.filename.split('.')[-1] if excel_file.filename else 'xlsx'
    with tempfile.NamedTemporaryFile(suffix=f".{excel_ext}", delete=False) as tmp_excel:
        excel_file.save(tmp_excel)
        tmp_excel_path = tmp_excel.name

    try:
        import pandas as pd

        # ── 엑셀 파싱: "개요서 번호" 컬럼 → target 순서 추출 ──
        excel_ext = Path(tmp_excel_path).suffix.lower()
        if excel_ext == '.csv':
            df_excel = pd.read_csv(tmp_excel_path, encoding='utf-8-sig')
        else:
            df_excel = pd.read_excel(tmp_excel_path, engine='openpyxl')

        cols = [str(c).strip() for c in df_excel.columns]

        # "개요서 번호" 컬럼 (새 순서 기준)
        def _find_col(keywords):
            for kw in keywords:
                for c in cols:
                    if kw in c:
                        return c
            return None

        # 엑셀에서 사용자 지정 컬럼 혹은 자동추론 
        usr_new_col = request.form.get("new_order_col")
        usr_old_col = request.form.get("old_order_col")

        if usr_new_col and usr_new_col in cols:
            new_order_col = usr_new_col
        else:
            new_order_col = _find_col(['개요서 번호', '개요서번호'])
            
        if usr_old_col and usr_old_col in cols:
            old_order_col = usr_old_col
        else:
            old_order_col = _find_col(['기존 번호', '기존번호', '이전 순번', '이전순번', '번호'])

        if new_order_col is None:
            # 개요서 번호가 없으면 load_reference_list 폴백
            ref_list = load_reference_list(tmp_excel_path)
            target_order_pairs = [(r.index, r.index) for r in ref_list]
        else:
            # (기존 PPT 번호, 엑셀 순서 기준 새 번호) 쌍 추출
            # old_order_col: PPT 슬라이드에서 찾을 before 번호
            # new_order_col: 재정렬 후 부여할 after 번호
            target_order_pairs = []
            for _, row in df_excel.iterrows():
                new_num_raw = row.get(new_order_col)
                if old_order_col:
                    old_num_raw = row.get(old_order_col)
                else:
                    old_num_raw = None
                try:
                    new_num = int(new_num_raw) if pd.notna(new_num_raw) else None
                except (ValueError, TypeError):
                    new_num = None
                try:
                    old_num = int(old_num_raw) if (old_num_raw is not None and pd.notna(old_num_raw)) else None
                except (ValueError, TypeError):
                    old_num = None

                if new_num is not None:
                    # old_num이 없으면 new_num을 PPT before 번호로도 사용
                    before_num = old_num if old_num is not None else new_num
                    target_order_pairs.append((before_num, new_num))

        # ── {before번호: after번호} 매핑 생성 ──
        mapping = {}
        for before_num, after_num in target_order_pairs:
            mapping[before_num] = after_num

        # 출력 파일 저장
        out_dir = session_store.get("output_dir", "output")
        Path(out_dir).mkdir(parents=True, exist_ok=True)
        output_filename = "output_briefs_reordered.pptx"
        output_path = str(Path(out_dir) / output_filename)

        # reorder_by_number_mapping: 레이블 번호 변경 + 물리적 순서 재배치를 한 번에 처리
        slide_count = reorder_by_number_mapping(tmp_ppt_path, mapping, output_path)

        return jsonify({
            "ok": True,
            "output_file": output_filename,
            "slide_count": slide_count,
            "mapping_used": {str(k): v for k, v in mapping.items()},
            "new_order_col": new_order_col,
            "old_order_col": old_order_col,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        os.unlink(tmp_ppt_path)
        os.unlink(tmp_excel_path)


@app.route("/api/download/<filename>", methods=["GET"])
def download_file(filename):
    out_dir = session_store.get("output_dir", "output")
    file_path = Path(out_dir) / filename
    if not file_path.exists():
        return jsonify({"error": "File not found"}), 404
    return send_file(str(file_path), as_attachment=True, download_name=filename)


# ---------- Generate Meta Excel (Case A: 엑셀 없음 → JSON에서 자동 생성) ----------
@app.route("/api/generate-meta-excel", methods=["POST"])
def generate_meta_excel():
    """현재 세션에서 처리한 JSON 파일들을 읽어 메타정보 엑셀을 생성합니다."""
    out_dir = Path(session_store.get("output_dir", "output"))

    # 현재 세션에서 생성된 JSON 파일 목록 (최신 job 기준)
    last_job = None
    for jb in reversed(list(session_store.get("jobs", {}).values())):
        if jb.get("json_files"):
            last_job = jb
            break

    if last_job and last_job.get("json_files"):
        # 세션에 기록된 파일만 읽기 (순번 정렬)
        json_paths = sorted(
            last_job["json_files"],
            key=lambda p: int(Path(p).stem) if Path(p).stem.isdigit() else float("inf")
        )
        json_files = [Path(p) for p in json_paths if Path(p).exists()]
        print(f"[INFO] generate_meta_excel: 세션 JSON {len(json_files)}개 사용")
    else:
        # 폴백: 세션 정보 없으면 out_dir 전체 (legacy)
        json_files = sorted(out_dir.glob("*.json"), key=lambda p: (
            int(p.stem) if p.stem.isdigit() else float("inf")
        ))
        print(f"[WARN] generate_meta_excel: 세션 정보 없음. 전체 JSON {len(json_files)}개 사용 (legacy 모드)")

    if not json_files:
        return jsonify({"error": "JSON 파일이 없습니다. 먼저 요약을 실행하세요."}), 400

    try:
        records = []
        for f in json_files:
            try:
                data = json.loads(f.read_text(encoding="utf-8"))
                data["_source_file"] = f.name
                records.append(data)
            except Exception:
                continue

        if not records:
            return jsonify({"error": "유효한 JSON 파일이 없습니다."}), 400

        # openpyxl로 엑셀 생성
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # match_results를 파일명/번호 기준으로 lookup 준비
        match_map = {}
        for mr in session_store.get("match_results", []):
            # source_pdf(원문 파일명) 기준으로 매핑
            if mr.get("filename"):
                match_map[mr["filename"]] = mr

        def flatten(data):
            pi = data.get("paper_info", {}) or {}
            def join(lst, sep="\n"):
                if not lst: return ""
                return sep.join(str(x).strip() for x in lst if x)
            def method_block(items):
                parts = []
                for m in (items or []):
                    parts.append(m.get("title", ""))
                    for d in m.get("details", []):
                        parts.append(f"  - {d.strip()}")
                return "\n".join(parts)
            # 대표 이미지 정보 추출
            rep_figs = data.get("representative_figures") or []
            fig1 = rep_figs[0] if len(rep_figs) > 0 else {}
            fig2 = rep_figs[1] if len(rep_figs) > 1 else {}
            caps_ko = data.get("figure_captions_ko") or ["", ""]
            cap1_ko = caps_ko[0] if len(caps_ko) > 0 else ""
            cap2_ko = caps_ko[1] if len(caps_ko) > 1 else ""
            # 문서유형 한글 변환
            raw_type = data.get("_meta_doc_type") or data.get("doc_type", "")
            doc_type_ko = {"patent": "특허", "paper": "논문"}.get(raw_type, raw_type)

            # 매칭 정보 병합
            source_pdf = data.get("source_pdf", "")
            mr = match_map.get(source_pdf, {})

            row = {
                "번호": data.get("_source_file", "").replace(".json", ""),
                "원문 파일명": source_pdf,
                "문서유형": doc_type_ko,
                "제목(기술명)": data.get("title", ""),
            }
            # 매칭 정보가 있으면 추가 (엑셀 제공 시)
            if match_map:
                row["매칭 기술명"] = mr.get("ref_title", "")
                row["일치도(%)"] = f"{mr.get('score', 0):.1f}" if mr.get("score") else ""
                row["매칭 상태"] = mr.get("status", "")

            row.update({
                "헤드메시지 1": (data.get("head_messages") or [""])[0],
                "헤드메시지 2": (data.get("head_messages") or ["", ""])[1] if len(data.get("head_messages") or []) > 1 else "",
                "기술목적": join(data.get("purpose", []), "\n"),
                "기존기술 문제점": join(data.get("prior_problems", []), "\n"),
                "제안기술": method_block(data.get("proposed_method", [])),
                "개선효과": join(data.get("improvements", []), "\n"),
                "저널/특허청": pi.get("journal_or_patent_office", ""),
                "논문/특허 제목(원문)": pi.get("paper_title", ""),
                "기관/출원인": pi.get("institution") or data.get("_meta_institution", ""),
                "DOI/특허번호": pi.get("doi_or_patent_no") or data.get("_meta_doi") or data.get("_meta_patent_no", ""),
                "연도": pi.get("year") or data.get("_meta_year", ""),
                "월": pi.get("month", ""),
                "대표이미지 번호1": f"Fig.{fig1.get('fig_number', '')}" if fig1.get('fig_number') else "",
                "대표이미지 캡션1": cap1_ko,
                "대표이미지 번호2": f"Fig.{fig2.get('fig_number', '')}" if fig2.get('fig_number') else "",
                "대표이미지 캡션2": cap2_ko,
            })
            return row

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "기술 개요 요약"
        rows = [flatten(r) for r in records]
        headers = list(rows[0].keys())

        THIN = Side(style="thin", color="B0BEC5")
        BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER

        for row_idx, row in enumerate(rows, 2):
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row.get(h, ""))
                cell.font = Font(size=9)
                cell.border = BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        ws.freeze_panes = "C2"
        ws.auto_filter.ref = ws.dimensions

        excel_filename = "briefs_summary.xlsx"
        excel_path = out_dir / excel_filename
        wb.save(str(excel_path))

        return jsonify({
            "ok": True,
            "output_file": excel_filename,
            "record_count": len(rows),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ---------- Reset ----------
@app.route("/api/reset", methods=["POST"])
def reset_session():
    session_store["documents"] = []
    session_store["order_data"] = []
    session_store["ref_list"] = []
    session_store["match_results"] = []
    session_store["jobs"] = {}
    return jsonify({"ok": True})


if __name__ == "__main__":
    import webbrowser
    def open_browser():
        time.sleep(1)  # 서버가 뜰 때까지 1초 대기
        webbrowser.open("http://localhost:5050")
        
    threading.Thread(target=open_browser, daemon=True).start()

    print("=" * 60)
    print("  후보기술 개요서 자동 요약 봇 — Flask API Server")
    print("  http://localhost:5050")
    print("=" * 60)
    app.run(host="0.0.0.0", port=5050, debug=True)
