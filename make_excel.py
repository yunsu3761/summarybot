#!/usr/bin/env python3
"""
output/*.json → Excel 변환 스크립트
각 JSON 파일의 메타정보를 행(row)으로 변환하여 엑셀로 저장
"""

import json
import os
import glob
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[오류] openpyxl이 설치되어 있지 않습니다. 설치 중...")
    os.system("pip install openpyxl")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path("output")
EXCEL_OUT  = OUTPUT_DIR / "briefs_summary.xlsx"


def load_all_jsons() -> list[dict]:
    """output/*.json을 번호 순서로 로드"""
    files = sorted(OUTPUT_DIR.glob("*.json"), key=lambda p: (
        int(p.stem) if p.stem.isdigit() else float("inf")
    ))
    records = []
    for f in files:
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
            data["_source_file"] = f.name
            records.append(data)
        except Exception as e:
            print(f"  [WARN] {f.name} 파싱 실패: {e}")
    return records


def flatten(data: dict) -> dict:
    """JSON dict를 엑셀 1행으로 펼치기"""
    pi = data.get("paper_info", {}) or {}

    def join(lst, sep="\n"):
        if not lst:
            return ""
        return sep.join(str(x).strip() for x in lst if x)

    def method_block(items):
        """proposed_method 배열 → 읽기 쉬운 텍스트"""
        parts = []
        for m in (items or []):
            title = m.get("title", "")
            details = m.get("details", [])
            parts.append(title)
            for d in details:
                parts.append(f"  - {d.strip()}")
        return "\n".join(parts)

    def improvements_text(lst):
        return join(lst, "\n")

    # 대표 이미지 정보 추출
    rep_figs = data.get("representative_figures") or []
    fig1 = rep_figs[0] if len(rep_figs) > 0 else {}
    fig2 = rep_figs[1] if len(rep_figs) > 1 else {}
    caps_ko = data.get("figure_captions_ko") or ["", ""]
    cap1_ko = caps_ko[0] if len(caps_ko) > 0 else ""
    cap2_ko = caps_ko[1] if len(caps_ko) > 1 else ""

    return {
        "번호":               data.get("_source_file", "").replace(".json", ""),
        "원문 파일명":         data.get("source_pdf", ""),
        "문서유형":           data.get("doc_type", ""),
        "제목(기술명)":       data.get("title", ""),
        "헤드메시지 1":       (data.get("head_messages") or [""])[0],
        "헤드메시지 2":       (data.get("head_messages") or ["", ""])[1] if len(data.get("head_messages") or []) > 1 else "",
        "기술목적":           join(data.get("purpose", []), "\n"),
        "기존기술 문제점":    join(data.get("prior_problems", []), "\n"),
        "제안기술":           method_block(data.get("proposed_method", [])),
        "개선효과":           improvements_text(data.get("improvements", [])),
        "저널/특허청":        pi.get("journal_or_patent_office", ""),
        "논문/특허 제목(원문)": pi.get("paper_title", ""),
        "기관/출원인":        pi.get("institution", ""),
        "DOI/특허번호":       pi.get("doi_or_patent_no", ""),
        "연도":               pi.get("year", ""),
        "월":                 pi.get("month", ""),
        "대표이미지 번호1":   f"Fig.{fig1.get('fig_number', '')}" if fig1.get('fig_number') else "",
        "대표이미지 캡션1":   cap1_ko,
        "대표이미지 번호2":   f"Fig.{fig2.get('fig_number', '')}" if fig2.get('fig_number') else "",
        "대표이미지 캡션2":   cap2_ko,
    }


# ── 스타일 상수 ────────────────────────────────────────────
HEADER_BG    = "1F4E79"   # 진한 파랑
HEADER_FG    = "FFFFFF"
PAPER_BG     = "EBF5FB"   # 연한 파랑 (논문)
PATENT_BG    = "FEF9E7"   # 연한 노랑 (특허)
ALT_PAPER    = "D6EAF8"
ALT_PATENT   = "FDF2D0"
BORDER_COLOR = "B0BEC5"

THIN = Side(style="thin", color=BORDER_COLOR)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 열 너비 (글자 단위)
COL_WIDTHS = {
    "번호":               6,
    "원문 파일명":         28,
    "문서유형":           8,
    "제목(기술명)":       40,
    "헤드메시지 1":       55,
    "헤드메시지 2":       55,
    "기술목적":           50,
    "기존기술 문제점":    50,
    "제안기술":           70,
    "개선효과":           50,
    "저널/특허청":        35,
    "논문/특허 제목(원문)": 55,
    "기관/출원인":        30,
    "DOI/특허번호":       35,
    "연도":               8,
    "월":                 8,
    "대표이미지 번호1":   14,
    "대표이미지 캡션1":   40,
    "대표이미지 번호2":   14,
    "대표이미지 캡션2":   40,
}


def build_excel(records: list[dict]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "기술 개요 요약"

    rows = [flatten(r) for r in records]
    if not rows:
        print("[오류] 변환할 데이터가 없습니다.")
        return

    headers = list(rows[0].keys())

    # ── 헤더 행 ──────────────────────────────────────────
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font      = Font(name="맑은 고딕", bold=True, color=HEADER_FG, size=10)
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(
            horizontal="center", vertical="center",
            wrap_text=True
        )
        cell.border = BORDER

    ws.row_dimensions[1].height = 30

    # ── 데이터 행 ────────────────────────────────────────
    for row_idx, row in enumerate(rows, 2):
        doc_type = row.get("문서유형", "")
        is_alt   = (row_idx % 2 == 0)

        if doc_type == "patent":
            bg = ALT_PATENT if is_alt else PATENT_BG
        else:
            bg = ALT_PAPER if is_alt else PAPER_BG

        fill = PatternFill("solid", fgColor=bg)

        for col_idx, h in enumerate(headers, 1):
            val  = row.get(h, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font  = Font(name="맑은 고딕", size=9)
            cell.fill  = fill
            cell.border= BORDER

            # 열별 정렬
            if h in ("번호", "문서유형", "연도", "월"):
                halign = "center"
            elif h in ("제안기술", "기술목적", "기존기술 문제점", "개선효과",
                       "헤드메시지 1", "헤드메시지 2"):
                halign = "left"
            else:
                halign = "left"

            cell.alignment = Alignment(
                horizontal=halign, vertical="top",
                wrap_text=True
            )

        ws.row_dimensions[row_idx].height = None  # 자동 높이

    # ── 열 너비 설정 ─────────────────────────────────────
    for col_idx, h in enumerate(headers, 1):
        width = COL_WIDTHS.get(h, 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── 틀 고정 (헤더 + 제목 열) ─────────────────────────
    ws.freeze_panes = "C2"

    # ── 자동 필터 ─────────────────────────────────────────
    ws.auto_filter.ref = ws.dimensions

    wb.save(EXCEL_OUT)
    print(f"\n✅ 엑셀 저장 완료: {EXCEL_OUT.resolve()}")
    print(f"   총 {len(rows)}개 기술 개요 포함")


if __name__ == "__main__":
    print("📂 JSON 파일 로딩 중...")
    records = load_all_jsons()
    print(f"  → {len(records)}개 JSON 로드 완료")
    print("📊 엑셀 생성 중...")
    build_excel(records)
